"""Rebuild the audit spreadsheet with a protein/biomarker column."""
from __future__ import annotations

import csv
import json
import glob
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

KNOWN_BIOMARKERS = re.compile(
    r"\b("
    r"NT-proBNP|pro-BNP|BNP|hs-CRP|hsCRP|sST2|ST2|Gal-3|galectin-3"
    r"|GDF-?15|IGFBP-?7|FGF-?21|IL-\d+|TNF-?\w*|TGF-?\w*|MMP-?\d*"
    r"|VEGF|PDGF|EGF|FGF|CTGF|TIMP-?\d*"
    r"|GLP-1|SGLT2i?|DPP-4|GIP"
    r"|NLR|PLR|SII|MPV|RDW|HbA1c"
    r"|CRP|SAA|PCT|ESR"
    r"|troponin[- ]?[TI]?|cTnT|cTnI|hs-?TnT|hs-?TnI"
    r"|albumin|globulin|ferritin|transferrin|hepcidin|cystatin[- ]?C?"
    r"|creatinine|urea|uric acid|homocysteine"
    r"|hemoglobin|myoglobin|copeptin|procalcitonin|presepsin"
    r"|natriuretic peptide|collagen|fibronectin|transthyretin|amyloid"
    r"|adiponectin|leptin|resistin|visfatin|apelin|ghrelin"
    r"|aldosterone|renin|angiotensin|endothelin"
    r"|cortisol|testosterone|estradiol|PTH|parathyroid hormone"
    r"|insulin|glucagon|calcitonin"
    r"|galectin|syndecan|interleukin|fibrinogen"
    r"|empagliflozin|dapagliflozin|canagliflozin|sotagliflozin"
    r"|sacubitril|valsartan|ivabradine|spironolactone|eplerenone"
    r"|sildenafil|tadalafil|riociguat|bosentan|macitentan"
    r"|alkaline phosphatase|bilirubin|AST|ALT|GGT"
    r"|QRS|T-wave|ST[- ]segment|PR interval"
    r"|Lp\(a\)|ApoB|LDL|HDL|VLDL|cholesterol|triglyceride"
    r")\b",
    re.IGNORECASE,
)

# Fallback: all-caps abbreviation with digit or 2+ uppercase letters
ABBREV_RE = re.compile(r"\b([A-Z]{2,7}(?:-\d{1,2})?)\b")
STOP_ABBREVS = {
    "HFpEF", "HFrEF", "HR", "CI", "OR", "BMI", "BP", "ECG", "EF", "LV", "RV",
    "LA", "RA", "CV", "CVD", "HF", "CHF", "AF", "NYHA", "ACC", "AHA", "ESC",
    "FDA", "RCT", "ICU", "MRI", "CT", "PET", "SPECT", "RNA", "DNA", "SD",
    "IQR", "SE", "ANOVA", "ROC", "AUC", "NRI", "IDI", "RESULTS", "METHODS",
    "CONCLUSIONS", "BACKGROUND", "OBJECTIVE", "PURPOSE", "DESIGN", "SETTING",
    "PATIENTS", "PARTICIPANTS", "DHF", "PH", "CAD", "CHD", "BB", "US", "UK",
    "CM", "OB", "ZSF", "DR", "QRS", "FT", "ATTR", "OF", "AS", "IS", "IN",
    "AN", "AT", "BY", "TO", "UP", "WE", "NO", "II", "VS", "OLS",
}


def extract_protein(sentence):
    """Extract human-readable protein/biomarker name from sentence text."""
    # First try known biomarker names
    m = KNOWN_BIOMARKERS.search(sentence)
    if m:
        return m.group(0)
    # Fallback: all-caps abbreviations
    for m in ABBREV_RE.finditer(sentence):
        if m.group(1) not in STOP_ABBREVS:
            return m.group(1)
    # Last resort: check for specific patterns
    special = {
        "FT3/FT4": "FT3/FT4 ratio",
        "ICAM1": "ICAM1",
        "myeloperoxidase": "myeloperoxidase",
        "platelet": "platelet function",
        "digoxin": "digoxin",
        "TyG": "TyG index",
        "ATTR-CM": "transthyretin amyloidosis",
        "ZSF1": "ZSF1 rat model",
        "diastolic dysfunction": "diastolic dysfunction",
    }
    for pattern, name in special.items():
        if pattern in sentence:
            return name
    return "(review needed)"


def main():
    with open("data/label_audit_v7_eval.json") as f:
        audit = json.load(f)
    with open("data/splits/hfpef_v7_eval_relabel3_noleak_large.json") as f:
        eval_data = json.load(f)

    flagged = [a for a in audit if a["majority_pred"] != a["gold_label"]]
    flagged.sort(key=lambda x: x["suspicion_score"], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relabeling Review"

    headers = [
        "Eval Index", "Protein/Biomarker", "Full Sentence", "Current Label",
        "Model Majority Vote", "Models Agreeing", "Total Models", "Suspicion Score",
        "P(associated)", "P(not_associated)", "P(incidental)",
        "New Label", "Notes",
    ]

    hfont = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    border = Border(bottom=Side(style="thin", color="D9D9D9"))
    label_colors = {
        "associated": PatternFill("solid", fgColor="FFA500"),
        "not_associated": PatternFill("solid", fgColor="92D050"),
        "incidental": PatternFill("solid", fgColor="FFD966"),
    }
    bfont = Font(name="Arial", size=9)
    bfont_bold = Font(name="Arial", size=9, bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, item in enumerate(flagged, 2):
        idx = item["index"]
        full_sentence = eval_data[idx]["sentence"]
        protein = extract_protein(full_sentence)

        values = [
            idx, protein, full_sentence, item["gold_label"], item["majority_pred"],
            item["n_agree_gold"], item["n_models"],
            round(item["suspicion_score"], 3),
            round(item["avg_probs"]["associated"], 3),
            round(item["avg_probs"]["not_associated"], 3),
            round(item["avg_probs"]["incidental"], 3),
            "", "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = bfont_bold if col == 2 else bfont
            cell.border = border
            if col == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col in (4, 5):
                cell.alignment = Alignment(horizontal="center")

        for c in (4, 5):
            label = values[c - 1]
            if label in label_colors:
                ws.cell(row=row_idx, column=c).fill = label_colors[label]

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 13
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 20

    dv = DataValidation(
        type="list", formula1='"associated,not_associated,incidental"', allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"L2:L{len(flagged) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(flagged) + 1}"

    out = Path("data/labeling/eval_relabel_audit_88.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    print(f"Saved {len(flagged)} rows to {out}")


if __name__ == "__main__":
    main()
