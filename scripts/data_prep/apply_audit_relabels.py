"""Apply relabel decisions from an audit workbook to a labeled JSON split."""
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


VALID_LABELS = {"associated", "not_associated", "incidental"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply audit workbook relabels")
    parser.add_argument("--sheet", required=True, help="Audit workbook path")
    parser.add_argument("--split", required=True, help="Input labeled JSON split")
    parser.add_argument("--index-column", required=True, help="Workbook column containing row indices")
    parser.add_argument("--output", required=True, help="Output labeled JSON split")
    parser.add_argument("--report", required=True, help="Output report JSON")
    parser.add_argument("--new-label-column", default="New Label")
    parser.add_argument("--current-label-column", default="Current Label")
    return parser.parse_args()


def load_relabels(
    sheet_path: Path,
    index_column: str,
    new_label_column: str,
    current_label_column: str,
) -> dict[int, dict[str, str]]:
    wb = load_workbook(sheet_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    header_idx = {header: i for i, header in enumerate(headers)}

    required = {index_column, new_label_column, current_label_column}
    missing = sorted(name for name in required if name not in header_idx)
    if missing:
        raise ValueError(f"Missing workbook columns: {missing}")

    relabels: dict[int, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_index = row[header_idx[index_column]]
        raw_new_label = row[header_idx[new_label_column]]
        raw_current_label = row[header_idx[current_label_column]]
        if raw_index is None or raw_new_label is None:
            continue

        index = int(raw_index)
        new_label = str(raw_new_label).strip()
        current_label = "" if raw_current_label is None else str(raw_current_label).strip()
        if new_label not in VALID_LABELS:
            raise ValueError(f"Invalid new label at row index {index}: {new_label}")
        if current_label and current_label not in VALID_LABELS:
            raise ValueError(f"Invalid current label at row index {index}: {current_label}")

        relabels[index] = {
            "current_label": current_label,
            "new_label": new_label,
        }
    return relabels


def main() -> None:
    args = parse_args()

    with open(args.split) as f:
        rows = json.load(f)

    relabels = load_relabels(
        sheet_path=Path(args.sheet),
        index_column=args.index_column,
        new_label_column=args.new_label_column,
        current_label_column=args.current_label_column,
    )

    updated_rows = list(rows)
    transitions = Counter()
    changed_indices: list[int] = []

    for index, relabel in sorted(relabels.items()):
        if index < 0 or index >= len(updated_rows):
            raise IndexError(f"Row index {index} is out of range for split of size {len(updated_rows)}")

        row = dict(updated_rows[index])
        split_label = row["label"]
        expected_label = relabel["current_label"]
        new_label = relabel["new_label"]

        if expected_label and split_label != expected_label:
            raise ValueError(
                f"Split label mismatch at index {index}: split={split_label} sheet={expected_label}"
            )

        if split_label == new_label:
            updated_rows[index] = row
            continue

        transitions[(split_label, new_label)] += 1
        changed_indices.append(index)
        row["label"] = new_label
        updated_rows[index] = row

    before = Counter(row["label"] for row in rows)
    after = Counter(row["label"] for row in updated_rows)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(updated_rows, f, indent=2)

    report = {
        "sheet": args.sheet,
        "split": args.split,
        "output": args.output,
        "total_rows": len(rows),
        "sheet_rows_with_new_label": len(relabels),
        "changed_count": len(changed_indices),
        "changed_indices": changed_indices,
        "before_distribution": dict(before),
        "after_distribution": dict(after),
        "transitions": {
            f"{old}->{new}": count for (old, new), count in sorted(transitions.items())
        },
    }

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w") as f:
        json.dump(report, f, indent=2)

    print("changed_count", report["changed_count"])
    print("before_distribution", report["before_distribution"])
    print("after_distribution", report["after_distribution"])
    print("saved_output", output_path)
    print("saved_report", report_path)


if __name__ == "__main__":
    main()
